import pandas as pd
from openpyxl  import Workbook, load_workbook
import math
wb = load_workbook('degerlendirmeler.xlsx')
sheet = wb.active

df_deg = pd.read_excel("degerlendirmeler.xlsx",header=1)
df_ogrnotlar = pd.read_excel("notlar.xlsx")
df_prgtablo = pd.read_excel("progcikti.xlsx",header=1)
ders_cikti_yuzdeleri =list(map(lambda x: float(x)/100,[cell.value for cell in sheet[1]][1:])) #ders cikti yuzdeleri okunuyor
ders_sayisi = len(ders_cikti_yuzdeleri)         

tablo3df =pd.DataFrame(columns=df_deg.columns.to_list()+["Toplam"]) #tablo3 dataframesinin sutunlari olusturuluyor
for x,deg_row in df_deg.iterrows(): #degerlendirmelerin tum satirlari okunuyor
    row = []
    row.append(deg_row.iloc[0])
    toplam = 0
    for i in range(1,1+ders_sayisi):
        deger = deg_row.iloc[i]* ders_cikti_yuzdeleri[i-1] #tum ders cikti degerleri ders yuzdeleri ile carpilip
        toplam+=deger                                      #degerlendirme sonucu elde ediliyor
        row.append(deger)
    row.append(toplam)
    tablo3df.loc[len(tablo3df)] = row


def truncate(number, digits):
    factor = 10 ** digits                               #bu fonksiyon verilen sayiyi digits sayisi kadar ,den sonra yazidiriyor
    return math.trunc(number * factor) / factor

tablo4_dfler={}  #tablo 4 dfleri Ogrenci adi key tablo4 ise item olacak sekilde ayarlaniyor cunku dosyaya daha rahat yazdirmak icin
for x,ogr_row in df_ogrnotlar.iterrows():   #her bir ogrenci notu icin
    ogr_adi = ogr_row.iloc[0]
    df = pd.DataFrame(columns=df_deg.columns.to_list()+["Toplam","MAX","% Başarı"])
    for y,ders_row in tablo3df.iterrows():    #tablo3 te kullanidigimz her bir agirlikli degerlendirme alinarak
        row = []
        row.append(ders_row.iloc[0])
        toplam = 0
        maxtoplam = 0
        for i in range(1,1+ders_sayisi):
            deger = ogr_row.iloc[i]*ders_row.iloc[i]        #ogrenci notlari ile hesaplaniyor
            maxdeger = ders_row.iloc[i]*100                 #maxdeger eldesi oran hesaplamak icin
            row.append(deger)
            maxtoplam += maxdeger
            toplam += deger
        row.append(toplam)
        row.append(maxtoplam)
        row.append(truncate(100*toplam/maxtoplam,2))                
        df.loc[len(df)] = row # yeni satir ekleniyor
    tablo4_dfler[ogr_adi] = df

tablo5_dfler = {} #tablo5 de tablo4de ki sebeplerden oturu ayni sekilde tanimlaniyor
#print(df_prgtablo.head())
for ogr_adi in tablo4_dfler.keys():    #her bir ogrenci icin
    basari_oranlari = tablo4_dfler[ogr_adi]["% Başarı"].tolist()                    #tablo 4 satirlari sutun olarak ayarlamiyor
    df = pd.DataFrame(columns=["Prg Çıktı"] + basari_oranlari + ["Başarı Oranı"])
    for x,prg_row in df_prgtablo.iterrows():    #her bir program ciktisi
        row = []
        toplam = 0
        maxtoplam= 0
        row.append(prg_row.iloc[0])
        for i in range(len(basari_oranlari)):
            deger = basari_oranlari[i] * prg_row.iloc[i+1] # basari orani ile carpilip hesaplaniyor
            maxtoplam +=(prg_row.iloc[i+1]*100)
            row.append(deger)
            toplam+= deger
        row.append(truncate(100*toplam/maxtoplam,2))
        df.loc[len(df)] = row #satira ekleniyor
    tablo5_dfler[ogr_adi] = df


# for i in tablo5_dfler.keys():
#     print("Ogrenci "+i + "\n",tablo5_dfler[i])

workbook = Workbook()
sheet = workbook.active
sheet.title = "Tablo3"

sheet["A1"] = "Tablo3"
sheet.merge_cells("B1:C1")  
sheet["B1"] = "Ağırlıklı Değerlendirme"

for c_idx, col_name in enumerate(tablo3df.columns, start=1):   ##tablo 3 1 satirdan itibaren yaziliyor
    sheet.cell(row=2, column=c_idx, value=col_name)

for r_idx, row in enumerate(tablo3df.itertuples(index=False), start=3):
    for c_idx, value in enumerate(row, start=1):
        sheet.cell(row=r_idx, column=c_idx, value=value)

workbook.save("tablo3.xlsx")


workbook = Workbook()
sheet = workbook.active
sheet.title = "Tablo4"

row_number= 1
for key in tablo4_dfler.keys():
    sheet.cell(row=row_number, column=1,value="Tablo4") #tablo 4 2satirdan itibaren yazilmaya basliyor satir satir
    sheet.merge_cells(start_row=row_number,end_row=row_number, start_column=2,end_column=3)
    sheet.cell(row=row_number,column=2,value="Ogrenci " + key + " için")
    row_number+=1
    for c_idx, col_name in enumerate(tablo4_dfler[key].columns, start=1):       #tablo4dfler dictimizdeki tum itemler sutun sutun satir satir yazdiriliyor
        sheet.cell(row=row_number, column=c_idx, value=col_name)

    for row in tablo4_dfler[key].itertuples(index=False):
        row_number+=1
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=c_idx, value=value)
    row_number+=2

workbook.save("tablo4.xlsx")


workbook = Workbook()
sheet = workbook.active
sheet.title = "Tablo4"
    
row_number= 1
for key in tablo5_dfler.keys():   #tablo 5 1.satirdan itibaren yazilmaya basliyor satir satir
    sheet.cell(row=row_number, column=1,value=" Ogrenci " + key + " için")
    row_number+=1
    sheet.cell(row=row_number, column=1, value="Tablo5")
    sheet.merge_cells(start_row=row_number,end_row=row_number, start_column=2,end_column=2+ders_sayisi)
    sheet.cell(row=row_number,column=2,value="Ders Çıktısı")
    row_number+=1   #tablo5dfler dictimizdeki tum itemler sutun sutun satir satir yazdiriliyor
    for c_idx, col_name in enumerate(tablo5_dfler[key].columns, start=1): 
        sheet.cell(row=row_number, column=c_idx, value=col_name)

    for row in tablo5_dfler[key].itertuples(index=False):
        row_number+=1
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=c_idx, value=value)
    row_number+=2
    
workbook.save("tablo5.xlsx")



